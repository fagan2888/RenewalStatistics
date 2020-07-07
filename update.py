import logging
import sqlite3

from openpyxl import load_workbook
# from datetime import datetime

logging.disable(logging.DEBUG)
logging.basicConfig(
    level=logging.DEBUG, format=" %(asctime)s | %(levelname)s | %(message)s"
)
logging.basicConfig(
    level=logging.INFO, format=" %(asctime)s | %(levelname)s | %(message)s"
)


def update(table_name):
    logging.info("开始更新数据库")
    data_path = r"Data\已续保清单.db"
    if table_name == "可续保清单":
        data_path = r"Data\可续保清单.db"
    conn = sqlite3.connect(data_path)
    logging.info("数据库连接成功")
    cur = conn.cursor()

    # 清空原数据库数据
    str_sql = f"DELETE FROM [{table_name}]"
    cur.execute(str_sql)
    conn.commit()
    logging.info("数据库数据清空完毕")

    # 读入Excel表格数据
    wb = load_workbook(f"{table_name}.xlsx")
    ws = wb["page"]

    max_col = ws.max_column
    max_row = ws.max_row

    logging.info("Excel 文件读入成功")
    logging.info(f"需要导入{ws.max_row}条数据")

    begin_row = 2

    if table_name == "可续保清单":
        for row in ws.iter_rows(
            min_row=begin_row, max_row=max_row, min_col=1, max_col=max_col
        ):
            str_sql = f"INSERT INTO '{table_name}' VALUES ( \
                '{row[0].value}', \
                '{row[1].value}', \
                '{row[2].value}', \
                '{row[3].value}', \
                '{row[4].value}', \
                '{row[5].value}', \
                '{row[6].value.strftime('%Y-%m-%d')}', \
                '{row[7].value.strftime('%Y-%m-%d')}', \
                '{row[8].value.strftime('%Y-%m-%d')}', \
                '{row[9].value}', \
                '{row[10].value}', \
                '{row[11].value}', \
                '{row[12].value}', \
                '{row[13].value}', \
                '{row[14].value}', \
                '{row[15].value}', \
                '{row[16].value}', \
                '{row[17].value}', \
                '{row[18].value}')"

            cur.execute(str_sql)

    else:
        for row in ws.iter_rows(
            min_row=begin_row, max_row=max_row, min_col=1, max_col=max_col
        ):
            str_sql = f"INSERT INTO '{table_name}' VALUES ( \
                '{row[0].value}', \
                '{row[1].value}', \
                '{row[2].value}', \
                '{row[3].value}', \
                '{row[4].value}', \
                '{row[5].value}', \
                '{row[6].value}', \
                '{row[7].value}', \
                '{row[8].value}', \
                '{row[9].value}', \
                '{row[10].value}', \
                '{row[11].value}', \
                '{row[12].value}', \
                '{row[13].value}', \
                '{row[14].value}', \
                '{row[15].value}')"

            cur.execute(str_sql)

    logging.info("数据写入数据库完成")

    conn.commit()

    logging.info("数据库事务提交完成")
    logging.info("数据库更新操作完成")

    cur.close()
    conn.close()

    print("-" * 60)


if __name__ == "__main__":
    # update("可续保清单")
    update("已续保清单")
